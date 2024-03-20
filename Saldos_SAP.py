import openpyxl
import win32com.client
import os
from tkinter import messagebox

#Função para saber o usuario da maquina
def usuario_maquina():

    full_username = os.getlogin()
    username_parts = full_username.split(" ")
    return username_parts[0] if len(username_parts) > 1 else username_parts[0]

#Definindo a Variável
username = usuario_maquina()

# #Conexão SAP
def conexao_SAP():
    SapGuiAuto = win32com.client.GetObject("SAPGUI")
    application = SapGuiAuto.GetScriptingEngine
    connection = application.Children(0)
    session = connection.Children(0)

    # Carrega o arquivo Excel
    pasta_trabalho = openpyxl.load_workbook('python_automacao_sap_saldos.xlsx')

    #Selecione a planilha
    pl_saldos = pasta_trabalho['Saldos']

    # Adicione esta linha no início do código
    linha_atual = 2

    #Maximizar a tela
    session.findById("wnd[0]").maximize()
    #Entrar na transação
    session.findById("wnd[0]/tbar[0]/okcd").Text = "S_ALR_87013019"
    #Confirmar entrada
    session.findById("wnd[0]").sendVKey(0)
    #Colocar na variante
    session.findById("wnd[0]/tbar[1]/btn[17]").press()
    #Selecionar a variante correta
    session.findById("wnd[1]/usr/cntlALV_CONTAINER_1/shellcont/shell").selectedRows = "0"
    #Corfimar e prosseguir
    session.findById("wnd[1]/tbar[0]/btn[2]").press()

    # Altere o loop para usar a variável linha_atual:
    for linhas in pl_saldos.iter_rows(min_row=linha_atual):

        if linha_atual > 1:
            # Inserir a ordem
            session.findById("wnd[0]/usr/ctxt_6ORDGRP-LOW").Text = pl_saldos.cell(row=linha_atual, column=1).value
            #Confirmar
            session.findById("wnd[0]").sendVKey(8)
            #Selecionar 2024
            session.findById("wnd[0]/shellcont/shell/shellcont[2]/shell").SelectedNode = "000007"
            # Maximizar a tela para pegar a posição correta de tela
            session.findById("wnd[0]").maximize()
            #Selecionar o Disponivél da ordem
            session.findById("wnd[0]/usr/lbl[110,8]").SetFocus()

            #Loop para cada linha
            nova_celula = pl_saldos.cell(row=linha_atual, column = 2)

            # atribuição
            nova_celula.value = session.findById("wnd[0]/usr/lbl[110,8]").Text
        
            session.findById("wnd[0]/tbar[0]/btn[3]").press()
            session.findById("wnd[1]/usr/btnBUTTON_YES").press()

        # Incremente a variável linha_atual após a verificação:
        linha_atual += 1
        print(nova_celula.value)
    # Salvar:
    pasta_trabalho.save('C:\\Users\\Wally\\Desktop\\vba\\Aut_Saldos.xlsx')

    print("Realizado com Sucesso!")
    messagebox.showinfo("Qualidade ICPO","Extração realizada!")
conexao_SAP()
